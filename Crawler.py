from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent

import re
import os
import openpyxl
from xl_helpx import move_to_
from test_soup import download_image
import time

import requests
from bs4 import BeautifulSoup
from urllib.request import Request, urlopen


class Crawler:
    def __init__(self, site_to_check, local_ss, listing_class_name):
        self.site_to_check = site_to_check

        options = Options()
        ua = UserAgent()
        user_agent = ua.random
        options.add_argument(f'user-agent={user_agent}')
        self.driver = webdriver.Chrome('./chromedriver.exe', chrome_options=options)

        self.wait = WebDriverWait(self.driver, 10)
        self.actions = ActionChains(self.driver)
        self.timeout = 20

        self.image_file = None

        self.change = False

        self.spread_filepath = os.path.join('C:/Users/colin/PycharmProjects/checker/sm_checker', 'items_data', local_ss)
        self.update_ss = os.path.join('C:/Users/colin/PycharmProjects/checker/sm_checker', 'items_data', 'master_upload.xlsx')
        self.wb = openpyxl.load_workbook(self.spread_filepath)
        self.ws = self.wb['master']
        self.wb2 = openpyxl.load_workbook(self.update_ss)

        self.driver.get(self.site_to_check)
        self.driver.maximize_window()

        self.wanted_items = []
        self.newRowLocation = self.ws.max_row + 1

        self.listing_class_name = listing_class_name

    def add_new_items(self, class_name, site_name, price_name, price_found=None):
        spreadsheet_items = self.load_spread_data()

        for item in self.wanted_items:
            self.click_item(item, spreadsheet_items, class_name, site_name, price_name, price_found)

    def load_spread_data(self):
        spreadsheet_items = []
        for row_num in range(2, self.ws.max_row + 1):
            spreadsheet_items.append(self.ws.cell(row=row_num, column=1).value)
        return spreadsheet_items

    def click_item(self, item, spreadsheet_items, class_name, site_name, price_name, price_found):
        if item not in spreadsheet_items:
            to_click = WebDriverWait(self.driver, self.timeout).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, item)))
            print(item)
            to_click.click()
            self.find_image_and_price(item, class_name, site_name, price_name, price_found)

    def find_image_and_price(self, item, class_name, site_name, price_name, price_found):
        image = self.get_image(class_name)
        price_e = self.get_price(price_name)

        price = self.check_price_found(price_found, price_e)
        print(image.get_attribute("src"))
        image_name = download_image(image.get_attribute("src"))
        self.update_row(item, 'None', 'Yes', image_name, 'to_upload', site_name, price)
        self.driver.back()
        self.change = True

    def check_price_found(self, price_found, price_element):
        if price_found:
            return price_element.get_attribute(price_found)
        else:
            print(price_element.text)
            return price_element.text

    def get_image(self, class_name):
        return WebDriverWait(self.driver, self.timeout).until(EC.element_to_be_clickable((By.CLASS_NAME, class_name)))

    def get_price(self, price_name):
        return self.driver.find_element_by_class_name(price_name)

    def update_row(self, series_name, style, in_stock, image_file, move_loc, site_name, price):
        short_name = self.get_short_name(series_name)
        size = self.get_size(series_name)
        self.ws.cell(column=1, row=self.newRowLocation, value=series_name)
        self.ws.cell(column=2, row=self.newRowLocation, value=short_name)
        self.ws.cell(column=3, row=self.newRowLocation, value=price)
        self.ws.cell(column=4, row=self.newRowLocation, value=size)
        self.ws.cell(column=5, row=self.newRowLocation, value=style)
        self.ws.cell(column=6, row=self.newRowLocation, value=in_stock)
        self.ws.cell(column=7, row=self.newRowLocation, value=self.driver.current_url)
        self.ws.cell(column=8, row=self.newRowLocation, value=image_file)
        self.ws.cell(column=9, row=self.newRowLocation, value=site_name)
        move_to_(self.newRowLocation, self.wb, self.wb2, move_loc)
        self.newRowLocation += 1

    def get_short_name(self, full_name):
        try:
            return full_name.split('"')[-1].split('-')[0]
        except AttributeError:
            return ' '

    def get_size(self, full_name):
        regexp = re.compile(r'"')
        regexp2 = re.compile(r"'")
        for x in full_name.split():
            if regexp.search(x):
                return x
            if regexp2.search(x):
                return x
        return 'N/A'

    # Out of stock to delete empty rows one big function
    def out_of_stock(self):
        # Check if item in spreadsheet but not on website
        spreadsheet_items = []
        for row_num in range(2, self.ws.max_row + 1):
            spreadsheet_items.append(self.ws.cell(row=row_num, column=1).value)

        self.check_for_series(spreadsheet_items)

    def check_for_series(self, spreadsheet_items):
        series_still_available = [series for series in self.wanted_items]
        print(series_still_available)
        out_of_stock_series = self.diff(series_still_available, spreadsheet_items)

        self.delete_oos_series(out_of_stock_series)

    def diff(self, li1, li2):
        return list(list(set(li1) - set(li2)) + list(set(li2) - set(li1)))

    def delete_oos_series(self, out_of_stock_series):
        # if item is still in the watchlist delete it from the spreadsheet
        rows_to_del = []
        for item in out_of_stock_series:
            for row in range(2, self.ws.max_row + 1):
                # for row in range(2, self.ws.max_row + 1):
                if self.ws.cell(row=row, column=1).value == item:
                    if self.ws.cell(row=row, column=3).value == 'Yes':
                        # Move to_delete sheet

                        move_to_(row, self.wb, self.wb2, 'to_delete')
                        self.change = True
                    self.ws.delete_rows(row)
                    self.ws.insert_rows(row)
                    print(f"{item} is no longer on the site. Deleting {item} entries from spreadhsheet")

                    # self.ws.delete_rows(row)

        self.delete_empty_rows()
        self.newRowLocation = self.ws.max_row + 1
        # ToDo Send email

    def delete_empty_rows(self):
        index_row = []

        # loop each row in column A
        for i in range(1, self.ws.max_row):
            # define emptiness of cell
            if self.ws.cell(i, 1).value is None:
                # collect indexes of rows
                index_row.append(i)

        # loop each index value
        for row_del in range(len(index_row)):
            self.ws.delete_rows(idx=index_row[row_del], amount=1)
            # exclude offset of rows through each iteration
            index_row = list(map(lambda k: k - 1, index_row))

    def close_workbooks(self):
        self.wb.save(self.spread_filepath)
        self.wb.close()
        self.wb2.save(self.update_ss)
        self.wb2.close()

    def load_items_from_site(self, class_name):
        # Add items on site to list
        items = self.driver.find_elements_by_class_name(class_name)
        for x in items:
            self.wanted_items.append(x.text)

    def run(self):
        self.load_items_from_site(self.listing_class_name)

        # check if an item is now out of stock
        self.out_of_stock()

        self.add_new_items(self.class_name, self.site_name, self.price_name, price_found=None)

        self.close_workbooks()
        self.driver.close()

        return self.change
