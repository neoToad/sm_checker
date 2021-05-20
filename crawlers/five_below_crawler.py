from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium
import time
import re

import openpyxl
from test_soup import find_styles, get_image
from xl_helpx import move_to_

import os


class FiveBelowCrawler:
    def __init__(self):
        self.site_to_check = "https://www.fivebelow.com/"

        self.driver = webdriver.Chrome('./chromedriver.exe')

        self.wait = WebDriverWait(self.driver, 10)
        self.actions = ActionChains(self.driver)

        # self.driver.get(self.site_to_check)

        self.image_file = None

        self.change = False

        self.spread_filepath = os.path.join('C:/Users/colin/PycharmProjects/checker', 'items_data', 'fb_sm.xlsx')
        self.wb = openpyxl.load_workbook(self.spread_filepath)

        print(self.spread_filepath)
        self.update_ss = os.path.join('C:/Users/colin/PycharmProjects/checker', 'items_data', 'master_upload.xlsx')

        self.ws = self.wb['master']
        self.wb2 = openpyxl.load_workbook(self.update_ss)


    def goto_stuffed_a(self):
        # open toys hover menu in top nav bar
        toys_menu = self.driver.find_element_by_link_text("toys & games")
        self.actions.move_to_element(toys_menu).perform()
        self.wait.until(EC.visibility_of_element_located((By.LINK_TEXT, "stuffed animals"))).click()

    def load_products(self):
        all_products_loading = True
        while all_products_loading:
            try:
                self.driver.find_element_by_id("loadMoreProducts").click()
            except selenium.common.exceptions.ElementNotInteractableException:
                all_products_loading = False

    def check_series_out_of_stock(self):
        # Check if item in spreadsheet but not on website
        spreadsheet_items = []
        for row_num in range(2, self.ws.max_row + 1):
            spreadsheet_items.append(self.ws.cell(row=row_num, column=1).value)

        self.check_for_series(spreadsheet_items)

    def check_for_series(self, spreadsheet_items):
        series_still_available = [series.text for series in self.wanted_items]
        out_of_stock_series = self.diff(series_still_available, spreadsheet_items)

        self.delete_oos_series(out_of_stock_series)

    def diff(self, li1, li2):
        return list(list(set(li1) - set(li2)) + list(set(li2) - set(li1)))

    def delete_oos_series(self, out_of_stock_series):
        # if item is still in the watchlist delete it from the spreadsheet
        rows_to_del = []
        for item in out_of_stock_series:
            for row in range(2, self.ws.max_row + 1):
                # for row in range(2, self.ws.max_row + 1):
                if self.ws.cell(row=row, column=1).value == item:
                    if self.ws.cell(row=row, column=3).value == 'Yes':
                        # Move to_delete sheet
                        move_to_(self.wb, self.wb2, 'to_delete')
                        self.change = True
                    self.ws.delete_rows(row)
                    self.ws.insert_rows(row)
                    print(f"{item} is no longer on the site. Deleting {item} entries from spreadhsheet")

                    # self.ws.delete_rows(row)

        self.delete_empty_rows()
        self.newRowLocation = self.ws.max_row + 1
        # ToDo Send email

    def delete_empty_rows(self):
        index_row = []

        # loop each row in column A
        for i in range(1, self.ws.max_row):
            # define emptiness of cell
            if self.ws.cell(i, 1).value is None:
                # collect indexes of rows
                index_row.append(i)

        # loop each index value
        for row_del in range(len(index_row)):
            self.ws.delete_rows(idx=index_row[row_del], amount=1)
            # exclude offset of rows through each iteration
            index_row = list(map(lambda k: k - 1, index_row))

    def known(self, item, style, in_stock):
        # Return true if website is current with spreadsheet.
        for row_num in range(2, self.ws.max_row + 1):
            item_series_name = self.ws.cell(row=row_num, column=1).value
            item_style = self.ws.cell(row=row_num, column=2).value
            item_instock = self.ws.cell(row=row_num, column=3).value
            if item_series_name == item and item_style == style and in_stock == item_instock:
                return True

    def check_for_items(self):
        for item in self.wanted_items:
            item.click()
            time.sleep(2)
            dif_styles = find_styles(self.driver.current_url)
            self.check_for_styles(dif_styles, item)

            self.driver.back()

    def check_for_styles(self, dif_styles, item):
        if dif_styles == False:  # If one style, check if known in spreadsheet and if not add it
            if not self.known(item.text, 'single', 'Yes'):
                self.image_file = get_image(self.driver.current_url, None, 'image-gallery-slide center')
                self.update_row(item.text, 'single', 'Yes', self.image_file, 'to_upload')
                self.change = True
        else:
            self.check_each_style(dif_styles, item)

    def check_each_style(self, styles, item):
        for style in styles:
            style_name_tag = self.driver.find_element_by_xpath(f"//*[name()='svg']/*[text()='{style}']")
            parent_tag = style_name_tag.find_element_by_xpath("./..")
            parent_tag.click()
            parent_cls_attr = parent_tag.get_attribute('class')
            parent_cls_attrs = parent_cls_attr.split()
            print(parent_cls_attrs)
            if 'active' in parent_cls_attr.split():
                print('Button active')
                if not self.known(item.text, style, 'Yes'):
                    self.image_file = get_image(self.driver.current_url, styles.index(style), 'image-gallery-slide '
                                                                                              'center')
                    self.update_row(item.text, style, 'Yes', self.image_file, 'to_upload')
                    self.change = True
                    print(f'{style} button is enabled')
            else:
                if not self.known(item.text, style, 'No'):
                    self.update_row(item.text, style, 'No', None, 'to_delete')
                    self.change = True
                    print(f'{style} button is NOT enabled')

    def update_row(self, series_name, style, in_stock, image_file, move_loc):
        # short_name = self.get_short_name(self.driver.current_url)
        self.ws.cell(column=1, row=self.newRowLocation, value=series_name)
        self.ws.cell(column=2, row=self.newRowLocation, value=style)
        self.ws.cell(column=3, row=self.newRowLocation, value=in_stock)
        self.ws.cell(column=4, row=self.newRowLocation, value=self.driver.current_url)
        self.ws.cell(column=5, row=self.newRowLocation, value=str(image_file))
        self.get_short_name(self.newRowLocation)
        move_to_(self.wb, self.wb2, move_loc)
        self.newRowLocation += 1

    def get_short_name(self, row):
        if self.ws.cell(column=6, row=row).value is None:
            long_url = self.ws.cell(column=4, row=row).value
            print(long_url)
            sn = re.search('https?://([A-Za-z_0-9.-]+).*', str(long_url))
            print(sn.group(1))
            self.ws.cell(column=6, row=row, value=sn.group(1)[4:])

    def add_to_spreadsheet(self, value, row, column):
        # TODO fit cell height to text
        self.ws.cell(column=column, row=row, value=value)

    def run_crawler(self):
        self.driver.get(self.site_to_check)
        self.driver.maximize_window()
        self.driver.find_element_by_tag_name('button').click()

        # wait = WebDriverWait(self.driver, 10)
        # actions = ActionChains(self.driver)
        try:
            self.goto_stuffed_a()
        except selenium.common.exceptions.NoSuchElementException:
            time.sleep(2)
            self.goto_stuffed_a()

        time.sleep(2)
        self.load_products()

        ## Find items with name

        self.wanted_items = self.driver.find_elements_by_partial_link_text("squishmallows")
        self.newRowLocation = self.ws.max_row + 1

        self.check_series_out_of_stock()
        self.check_for_items()

        self.wb.save(self.spread_filepath)
        self.wb.close()
        self.wb2.save(self.update_ss)
        self.wb2.close()
        self.driver.quit()
        return self.change


# crawler = FiveBelowCrawler()
#
# crawler.run_crawler()
