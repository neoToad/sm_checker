import openpyxl
import os


def move_to_(row, wb, wb2, destination):

    ws1 = wb['master']
    ws2 = wb2[destination]
    mc = ws1.max_column
    mr = ws2.max_row
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=row, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=mr + 1, column=j).value = c.value

    # saving the destination excel file
    # wb.save(filename='fb_items_data/master_upload.xlsx')


# move_to_(2, openpyxl.load_workbook('fb_items_data/master_upload.xlsx'), 'to_upload')


def delete_updated_items_ss():
    ss_file_path = os.path.join('C:/Users/colin/PycharmProjects/checker/sm_checker', 'items_data', 'master_upload.xlsx')
    wb = openpyxl.load_workbook(ss_file_path)
    ws = wb['to_upload']
    while ws.max_row > 1:
        # this method removes the row 2
        ws.delete_rows(2)
        # return to main function

    ws = wb['to_delete']
    while ws.max_row > 1:
        # this method removes the row 2
        ws.delete_rows(2)
        # return to main function

    wb.save(ss_file_path)
    wb.close()


